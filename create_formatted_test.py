import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook

def create_formatted_test():
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active

        # Define data
        data = [
            ['学校名称', '专业名称', '学校代码', '专业代码', '位次差', '招生类型', '办学性质', '专业类', '24选科', '省份', '城市', '学费'],
            ['清华大学', '计算机科学', '10001', '0801', '+50', '综合', '公办', '工学', '物理', '北京', '海淀', '5000'],
            ['北京大学', '软件工程', '10002', '0802', '-30', '理科', '公办', '工学', '物理', '北京', '海淀', '5500'],
            ['浙江大学', '人工智能', '10003', '0803', '+20', '综合', '公办', '工学', '物理', '浙江', '杭州', '6000']
        ]

        # Write data and apply formatting
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Apply formatting based on row
                if row_idx == 2:  # First data row - Blue text
                    cell.font = Font(color='0000FF')
                elif row_idx == 3:  # Second data row - Red background
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                elif row_idx == 4:  # Third data row - Purple background
                    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

        # Save the file
        wb.save('formatted_test.xlsx')
        print("Successfully created formatted test file with:")
        print("- Row 1: Blue text (清华大学)")
        print("- Row 2: Red background (北京大学)")
        print("- Row 3: Purple background (浙江大学)")
        return True
    except Exception as e:
        print(f"Error creating test file: {str(e)}")
        return False

if __name__ == "__main__":
    create_formatted_test()
