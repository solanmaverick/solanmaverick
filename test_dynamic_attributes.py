import openpyxl
from openpyxl.styles import PatternFill, Font
import os

def create_test_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Headers - intentionally missing some fixed attributes to test parser
    headers = ['学校名称', '专业名称', '新属性1', '新属性2']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Add some test data
    data = [
        ['清华大学', '计算机科学', '值1', '值2'],
        ['北京大学', '软件工程', '值3', '值4']
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 2:
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color='FFFF0000',
                    end_color='FFFF0000',
                    fill_type='solid'
                )

    # Save the file in the current directory
    filename = os.path.join(os.path.dirname(__file__), 'test_dynamic.xlsx')
    wb.save(filename)
    return filename

def verify_file(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    print("Excel File Contents:")
    for row in ws.iter_rows(values_only=True):
        print(row)

    print("\nHeaders found:", [cell.value for cell in ws[1]])

if __name__ == '__main__':
    filename = create_test_file()
    verify_file(filename)
