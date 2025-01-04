import openpyxl
from openpyxl.styles import PatternFill, Font, Color, Alignment
import random
import os

def create_test_file(filename, scenario):
    """Create test Excel files for different scenarios."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Ensure the output directory exists
    os.makedirs('test_files', exist_ok=True)
    filename = os.path.join('test_files', filename)
    
    # Fixed attributes that must be present
    fixed_attrs = ['学校名称', '专业名称', '学校代码', '专业代码', '位次差']
    
    # Dynamic attributes for testing
    dynamic_attrs = ['录取批次', '录取年份', '录取分数', '录取人数', '专业类别', '学费', '就业率']
    
    if scenario == 'random_positions':
        # Test random positions of fixed attributes
        all_attrs = fixed_attrs + dynamic_attrs
        random.shuffle(all_attrs)
        headers = all_attrs
        
    elif scenario == 'missing_values':
        # Test with empty fixed attributes
        headers = fixed_attrs + dynamic_attrs
        
    elif scenario == 'dynamic_only':
        # Test with only dynamic attributes first, then add fixed ones
        headers = dynamic_attrs
        
    elif scenario == 'styled_cells':
        # Test with cell styling
        headers = fixed_attrs + dynamic_attrs
        
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='000000')  # Black color for headers
        if scenario == 'styled_cells':
            if header in ['学校名称', '专业名称']:
                cell.font = Font(bold=True, color='0000FF')  # Blue
            elif header in ['学校代码', '专业代码']:
                cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            elif header == '位次差':
                cell.font = Font(bold=True, color='FF0000')  # Red
    
    # Sample data
    sample_data = {
        '学校名称': ['清华大学', '北京大学', '浙江大学'],
        '专业名称': ['计算机科学与技术', '软件工程', '人工智能'],
        '学校代码': ['10003', '10001', '10335'],
        '专业代码': ['0809', '0810', '0854'],
        '位次差': ['+50', '-30', '+20'],
        '录取批次': ['提前批', '一批', '一批'],
        '录取年份': ['2023', '2023', '2023'],
        '录取分数': ['680', '675', '670'],
        '录取人数': ['120', '100', '150'],
        '专业类别': ['工学', '工学', '工学'],
        '学费': ['5800', '5800', '5500'],
        '就业率': ['98.5%', '97.8%', '96.5%']
    }
    
    # Write data rows
    for row in range(2, 5):
        for col, header in enumerate(headers, 1):
            if header in sample_data:
                value = sample_data[header][row-2]
                cell = ws.cell(row=row, column=col, value=value)
                
                if scenario == 'styled_cells':
                    # Add explicit styling for all cells
                    if header in ['学校名称', '专业名称']:
                        cell.font = Font(color='0000FF')  # Blue text
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    elif header in ['学校代码', '专业代码']:
                        cell.font = Font(name='Consolas', color='000000')  # Monospace font
                        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    elif header == '位次差':
                        if str(value).startswith('+'):
                            cell.font = Font(color='008000', bold=True)  # Green bold for positive
                            cell.fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
                        else:
                            cell.font = Font(color='FF0000', bold=True)  # Red bold for negative
                            cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    else:
                        # Style for dynamic attributes
                        cell.font = Font(color='000000')  # Black text
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            else:
                # Empty cell for missing values scenario
                ws.cell(row=row, column=col, value='')
    
    wb.save(filename)

# Create test files for different scenarios
scenarios = ['random_positions', 'missing_values', 'dynamic_only', 'styled_cells']
for scenario in scenarios:
    create_test_file(f'test_{scenario}.xlsx', scenario)
